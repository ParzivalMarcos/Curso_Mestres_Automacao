from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import openpyxl
import os

class BuscarValor():
    
    def __init__(self):
        chrome_options = Options()
        chrome_options.add_argument('--lang=pt-BR')
        self.driver = webdriver.Chrome(
            executable_path=os.getcwd() + os.sep + 'chromedriver.exe'
        )

    
    def Iniciar(self):
        self.numero_pagina = 1
        self.driver.get(f'https://sp.olx.com.br/sao-paulo-e-regiao/centro?o={self.numero_pagina}&q=bicicleta')
        self.criar_planilha()
        self.capturar_informacoes_do_site()

    
    def criar_planilha(self):
        self.planilha = openpyxl.Workbook()
        self.guia_planilha = self.planilha.create_sheet('Valores da OLX')
        self.guia_planilha.cell(row=1, column=1, value='Titulo')
        self.guia_planilha.cell(row=1, column=2, value='Localidade')
        self.guia_planilha.cell(row=1, column=3, value='Preço')
        # planilha.save('Projeto_WEB_excel.xlsx')


    def capturar_informacoes_do_site(self):
        
    # try:
        while True:
            self.titulos = self.driver.find_elements_by_xpath('//*[@class="fnmrjs-6 iNpuEh"]')
            self.localidades = self.driver.find_elements_by_xpath('//*[@class="sc-ifAKCX sc-7l84qu-1 fVmnUX"]')
            self.preco = self.driver.find_elements_by_xpath('//*[@class="fnmrjs-9 dCffJM"]')

            if not self.titulos:
                print('Pagina não encontrada !!')
                self.driver.close()
                break

            self.armazenar_dados_na_planilha()
            self.navegar_para_proxima_pagina()

    # except Exception as erro:
    #     print('Não existe mais paginas !!')
    #     print(erro)


    def armazenar_dados_na_planilha(self):
        for indice in range(0, len(self.titulos)):
            linha = [self.titulos[indice].text,
                     self.localidades[indice].text,
                     self.preco[indice].text]
            self.guia_planilha.append(linha)
        self.planilha.save('Preços_Bicicletas_centro_SP.xlsx')


    def navegar_para_proxima_pagina(self):
        self.numero_pagina += 1
        self.driver.get(f'https://sp.olx.com.br/sao-paulo-e-regiao/centro?o={self.numero_pagina}&q=bicicleta')


busca = BuscarValor()
busca.Iniciar()
