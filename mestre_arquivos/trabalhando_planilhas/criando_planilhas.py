import openpyxl

planilha_test = openpyxl.Workbook()

planilha_test.create_sheet('Frutas')
planilha_test.create_sheet('Legumes')
planilha_test.create_sheet('Sementes')
print(planilha_test.sheetnames)
cabecalho_frutas = ['Título', 'Localização', 'Preço']

planilha_frutas = planilha_test['Frutas']
planilha_frutas.append(cabecalho_frutas)
planilha_frutas.title = 'Frutas em Promoção'
planilha_test.save('Dados Supermercado.xlsx')
