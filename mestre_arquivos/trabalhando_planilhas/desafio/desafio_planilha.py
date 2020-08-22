import openpyxl

planilha = openpyxl.load_workbook('ExemploPlanilha.xlsx')

registros2015 = planilha['Registros 2015']

# Imprima o valor Fallon
print(registros2015['B11'].value)

# Altere o valor dessa celula para 'Falcon'
# 1° Modo
registros2015['B11'] = 'Falcon'
print(registros2015['B11'].value)

# 2º Modo
registros2015.cell(row=11, column=2, value='Falcon')
print(registros2015['B11'].value)

# Imprima na tela somente os valores das linhas 2 a 10
for linha in registros2015.iter_rows(min_row=2, max_row=10, min_col=1, max_col=None):
    for celula in linha:
        print(celula.value)

# Imprima na tela todos os valores da coluna 'Last Name'
for coluna in registros2015.iter_cols(min_col=2, max_col=2, min_row=2, max_row=None):
    for celula in coluna:
        print(celula.value)