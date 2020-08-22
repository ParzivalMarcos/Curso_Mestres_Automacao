import openpyxl

planilha = openpyxl.Workbook()

planilha.create_sheet("Preços de Canetas")
planilha_canetas = planilha['Preços de Canetas']

planilha_canetas['A1'] = '=SUM(1,1)'
planilha_canetas['A2'] = '=SUM(5,5)'
planilha_canetas['A3'] = '=SUM(10,10)'
planilha_canetas['A4'] = 'AVERAGE(A1:A3)'
planilha_canetas['A5'] = 'MIN(A1:A4)'
planilha_canetas['A6'] = 'MAX(A1:A4)'

# Sintaxe alternativa
planilha_canetas.cell(row=1, column=1, value='=SUM(1,1)')
planilha_canetas.cell(row=2, column=1, value='=SUM(5,5)')
planilha_canetas.cell(row=3, column=1, value='=SUM(10,10)')
planilha_canetas.cell(row=4, column=1, value='=AVERAGE(A1:A3)')
planilha_canetas.cell(row=5, column=1, value='=MIN(A1:A4)')
planilha_canetas.cell(row=6, column=1, value='=MAX(A1:A4)')

planilha.save('Planilha Canetas.xlsx')