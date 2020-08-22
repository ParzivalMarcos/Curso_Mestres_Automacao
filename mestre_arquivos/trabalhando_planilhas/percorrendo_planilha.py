import openpyxl

# Sheet - Guia da planilha
# Workbook - Planilha como um todo

planilha = openpyxl.load_workbook('ExemploPlanilha.xlsx')
print(planilha.sheetnames)

# Selecionando uma guia e mostrando o valor de uma celula
planilha_registros2015 = planilha['Registros 2015']
print(planilha_registros2015['B2'].value)

# Alterando o valor de uma celula
planilha_registros2015['B2'] = 'Dulce Gusto'

# Salvando
planilha.save('ExemploPlanilhaV2.xlsx')

# Outra forma de alterar valor de uma celula
planilha_registros2015.cell(row=2, column=2, value=5)

# Iterando uma planilha
for linha in planilha_registros2015.iter_rows(min_row=2, max_row=None, min_col=1, max_col=8):
    for celula in linha:
        print(celula.value)

for coluna in planilha_registros2015.iter_cols(min_col=3, max_col=3, min_row=2, max_row=None):
    for celula in coluna:
        print(celula.value)